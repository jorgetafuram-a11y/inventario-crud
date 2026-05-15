from datetime import datetime, timedelta
from io import BytesIO
from functools import wraps
from secrets import choice

from flask import ( # type: ignore
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
 

app = Flask(__name__)
app.config["SECRET_KEY"] = "1010115617"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///inventario.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    recovery_code = db.Column(db.String(8), nullable=True)
    recovery_expires_at = db.Column(db.DateTime, nullable=True)


class Part(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(40), unique=True, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    vehicle_type = db.Column(db.String(20), nullable=False)  # carro | moto
    brand = db.Column(db.String(80), nullable=True)
    stock = db.Column(db.Integer, nullable=False, default=0)
    price = db.Column(db.Float, nullable=False, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Invoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_name = db.Column(db.String(120), nullable=False)
    total = db.Column(db.Float, nullable=False, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)

    created_by = db.relationship("User", backref="invoices")


class InvoiceItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey("invoice.id"), nullable=False)
    part_id = db.Column(db.Integer, db.ForeignKey("part.id"), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)

    invoice = db.relationship("Invoice", backref="items")
    part = db.relationship("Part")


def seed_users() -> None:
    if not User.query.filter_by(username="superadmin").first():
        admin = User(
            username="superadmin",
            password_hash=generate_password_hash("Admin123*"),
            role="super_admin",
        )
        db.session.add(admin)

    if not User.query.filter_by(username="personal").first():
        staff = User(
            username="personal",
            password_hash=generate_password_hash("Personal123*"),
            role="personal",
        )
        db.session.add(staff)

    db.session.commit()


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            flash("Debes iniciar sesion.", "warning")
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapper


def role_required(*roles):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(*args, **kwargs):
            if session.get("role") not in roles:
                flash("No tienes permisos para esta accion.", "danger")
                return redirect(url_for("dashboard"))
            return view_func(*args, **kwargs)

        return wrapper

    return decorator


def get_current_user() -> User | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    return db.session.get(User, user_id)


def ensure_user_recovery_columns() -> None:
    columns = {
        row[1]
        for row in db.session.execute(db.text("PRAGMA table_info(user)")).fetchall()
    }

    if "recovery_code" not in columns:
        db.session.execute(
            db.text("ALTER TABLE user ADD COLUMN recovery_code VARCHAR(8)")
        )
    if "recovery_expires_at" not in columns:
        db.session.execute(
            db.text("ALTER TABLE user ADD COLUMN recovery_expires_at DATETIME")
        )

    db.session.commit()


def generate_password_recovery_code() -> str:
    digits = "0123456789"
    while True:
        code = "".join(choice(digits) for _ in range(8))
        exists = User.query.filter_by(recovery_code=code).first()
        if not exists:
            return code


def validate_password_recovery_code(code: str) -> User | None:
    user = User.query.filter_by(recovery_code=code).first()
    if not user or not user.recovery_expires_at:
        return None

    if user.recovery_expires_at < datetime.utcnow():
        user.recovery_code = None
        user.recovery_expires_at = None
        db.session.commit()
        return None

    return user


@app.context_processor
def inject_user():
    return {"current_user": get_current_user()}


@app.route("/")
def root():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = User.query.filter_by(username=username).first()
        if not user or not check_password_hash(user.password_hash, password):
            flash("Usuario o contrasena invalida.", "danger")
            return render_template("login.html")

        session["user_id"] = user.id
        session["role"] = user.role
        flash("Sesion iniciada correctamente.", "success")
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.route("/registro", methods=["GET", "POST"])
def register():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        password_confirm = request.form.get("password_confirm", "")

        # Validaciones
        if not username or not password:
            flash("Usuario y contraseña son requeridos.", "danger")
            return render_template("register.html")

        if len(username) < 3:
            flash("El usuario debe tener al menos 3 caracteres.", "danger")
            return render_template("register.html", username=username)

        if len(password) < 8:
            flash("La contraseña debe tener al menos 8 caracteres.", "danger")
            return render_template("register.html", username=username)

        if password != password_confirm:
            flash("Las contraseñas no coinciden.", "danger")
            return render_template("register.html", username=username)

        if User.query.filter_by(username=username).first():
            flash("El usuario ya existe.", "danger")
            return render_template("register.html")

        # Crear nuevo usuario con rol "personal"
        user = User(
            username=username,
            password_hash=generate_password_hash(password),
            role="personal",
        )
        db.session.add(user)
        db.session.commit()
        
        flash("Cuenta creada exitosamente. Ahora puedes iniciar sesión.", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


@app.route("/logout")
@login_required
def logout():
    session.clear()
    flash("Sesion cerrada.", "info")
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    parts_count = Part.query.count()
    low_stock = Part.query.filter(Part.stock <= 5).count()
    invoices_count = Invoice.query.count()
    total_sales = db.session.query(db.func.coalesce(db.func.sum(Invoice.total), 0)).scalar()

    return render_template(
        "dashboard.html",
        parts_count=parts_count,
        low_stock=low_stock,
        invoices_count=invoices_count,
        total_sales=total_sales,
    )


@app.route("/inventario")
@login_required
def inventory_list():
    search = request.args.get("search", "").strip()
    vehicle = request.args.get("vehicle", "").strip()

    query = Part.query
    if search:
        like_value = f"%{search}%"
        query = query.filter(
            db.or_(
                Part.code.ilike(like_value),
                Part.name.ilike(like_value),
                Part.brand.ilike(like_value),
            )
        )
    if vehicle in {"carro", "moto"}:
        query = query.filter_by(vehicle_type=vehicle)

    parts = query.order_by(Part.created_at.desc()).all()
    return render_template("inventory_list.html", parts=parts, search=search, vehicle=vehicle)


@app.route("/inventario/nuevo", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "personal")
def inventory_create():
    if request.method == "POST":
        code = request.form.get("code", "").strip().upper()
        name = request.form.get("name", "").strip()
        vehicle_type = request.form.get("vehicle_type", "").strip().lower()
        brand = request.form.get("brand", "").strip()
        stock = request.form.get("stock", "0").strip()
        price = request.form.get("price", "0").strip()

        if not code or not name or vehicle_type not in {"carro", "moto"}:
            flash("Completa los campos obligatorios.", "danger")
            return render_template("inventory_form.html", part=None)

        if Part.query.filter_by(code=code).first():
            flash("Ya existe un repuesto con ese codigo.", "danger")
            return render_template("inventory_form.html", part=None)

        try:
            stock_val = max(0, int(stock))
            price_val = max(0.0, float(price))
        except ValueError:
            flash("Stock o precio invalido.", "danger")
            return render_template("inventory_form.html", part=None)

        new_part = Part(
            code=code,
            name=name,
            vehicle_type=vehicle_type,
            brand=brand,
            stock=stock_val,
            price=price_val,
        )
        db.session.add(new_part)
        db.session.commit()
        flash("Repuesto creado correctamente.", "success")
        return redirect(url_for("inventory_list"))

    return render_template("inventory_form.html", part=None)


@app.route("/inventario/<int:part_id>/editar", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "personal")
def inventory_edit(part_id: int):
    part = db.session.get(Part, part_id)
    if not part:
        flash("Repuesto no encontrado.", "danger")
        return redirect(url_for("inventory_list"))

    if request.method == "POST":
        code = request.form.get("code", "").strip().upper()
        name = request.form.get("name", "").strip()
        vehicle_type = request.form.get("vehicle_type", "").strip().lower()
        brand = request.form.get("brand", "").strip()
        stock = request.form.get("stock", "0").strip()
        price = request.form.get("price", "0").strip()

        if not code or not name or vehicle_type not in {"carro", "moto"}:
            flash("Completa los campos obligatorios.", "danger")
            return render_template("inventory_form.html", part=part)

        code_exists = Part.query.filter(Part.code == code, Part.id != part.id).first()
        if code_exists:
            flash("El codigo ya esta en uso por otro repuesto.", "danger")
            return render_template("inventory_form.html", part=part)

        try:
            stock_val = max(0, int(stock))
            price_val = max(0.0, float(price))
        except ValueError:
            flash("Stock o precio invalido.", "danger")
            return render_template("inventory_form.html", part=part)

        part.code = code
        part.name = name
        part.vehicle_type = vehicle_type
        part.brand = brand
        part.stock = stock_val
        part.price = price_val

        db.session.commit()
        flash("Repuesto actualizado correctamente.", "success")
        return redirect(url_for("inventory_list"))

    return render_template("inventory_form.html", part=part)


@app.route("/inventario/<int:part_id>/eliminar", methods=["POST"])
@login_required
@role_required("super_admin")
def inventory_delete(part_id: int):
    part = db.session.get(Part, part_id)
    if not part:
        flash("Repuesto no encontrado.", "danger")
        return redirect(url_for("inventory_list"))

    db.session.delete(part)
    db.session.commit()
    flash("Repuesto eliminado.", "info")
    return redirect(url_for("inventory_list"))


@app.route("/facturas")
@login_required
def invoice_list():
    invoices = Invoice.query.order_by(Invoice.created_at.desc()).all()
    return render_template("invoice_list.html", invoices=invoices)


@app.route("/facturas/nueva", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "personal")
def invoice_create():
    parts = Part.query.order_by(Part.name.asc()).all()

    if request.method == "POST":
        customer_name = request.form.get("customer_name", "").strip()
        if not customer_name:
            flash("El nombre del cliente es obligatorio.", "danger")
            return render_template("invoice_form.html", parts=parts)

        part_ids = request.form.getlist("part_id")
        quantities = request.form.getlist("quantity")

        selected_items = []
        total = 0.0

        for raw_part_id, raw_quantity in zip(part_ids, quantities):
            try:
                quantity = int(raw_quantity)
                part_id = int(raw_part_id)
            except ValueError:
                continue

            if quantity <= 0:
                continue

            part = db.session.get(Part, part_id)
            if not part:
                continue

            if quantity > part.stock:
                flash(f"Stock insuficiente para {part.name}.", "danger")
                return render_template("invoice_form.html", parts=parts)

            subtotal = quantity * part.price
            total += subtotal
            selected_items.append((part, quantity, part.price, subtotal))

        if not selected_items:
            flash("Debes agregar al menos un repuesto con cantidad mayor a 0.", "danger")
            return render_template("invoice_form.html", parts=parts)

        invoice = Invoice(
            customer_name=customer_name,
            total=total,
            created_by_id=session["user_id"],
        )
        db.session.add(invoice)
        db.session.flush()

        for part, quantity, unit_price, subtotal in selected_items:
            item = InvoiceItem(
                invoice_id=invoice.id,
                part_id=part.id,
                quantity=quantity,
                unit_price=unit_price,
                subtotal=subtotal,
            )
            part.stock -= quantity
            db.session.add(item)

        db.session.commit()
        flash("Factura creada correctamente.", "success")
        return redirect(url_for("invoice_list"))

    return render_template("invoice_form.html", parts=parts)


@app.route("/facturas/<int:invoice_id>/excel")
@login_required
def invoice_export_excel(invoice_id: int):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash("Factura no encontrada.", "danger")
        return redirect(url_for("invoice_list"))

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Factura"

    ws["A1"] = "Factura"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:E1")

    ws["A2"] = f"Numero: {invoice.id}"
    ws["A3"] = f"Cliente: {invoice.customer_name}"
    ws["A4"] = f"Fecha: {invoice.created_at.strftime('%Y-%m-%d %H:%M')}"
    ws["A5"] = f"Vendedor: {invoice.created_by.username}"

    headers = ["Codigo", "Nombre", "Cantidad", "Precio Unitario", "Subtotal"]
    header_row = 7
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row_idx = header_row + 1
    for item in invoice.items:
        ws.cell(row=row_idx, column=1, value=item.part.code)
        ws.cell(row=row_idx, column=2, value=item.part.name)
        ws.cell(row=row_idx, column=3, value=item.quantity)
        ws.cell(row=row_idx, column=4, value=item.unit_price)
        ws.cell(row=row_idx, column=5, value=item.subtotal)
        row_idx += 1

    total_row = row_idx + 1
    ws.cell(row=total_row, column=4, value="Total").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=invoice.total).font = Font(bold=True)

    last_data_row = row_idx - 1
    if last_data_row >= header_row + 1:
        table_ref = f"A{header_row}:E{last_data_row}"
        table = Table(displayName=f"Factura{invoice.id}", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"factura_{invoice.id}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/facturas/<int:invoice_id>/eliminar", methods=["POST"])
@login_required
@role_required("super_admin")
def invoice_delete(invoice_id: int):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash("Factura no encontrada.", "danger")
        return redirect(url_for("invoice_list"))

    for item in invoice.items:
        if item.part:
            item.part.stock += item.quantity

    InvoiceItem.query.filter_by(invoice_id=invoice.id).delete()
    db.session.delete(invoice)
    db.session.commit()
    flash("Factura eliminada correctamente y stock restaurado.", "info")
    return redirect(url_for("invoice_list"))


@app.route("/usuarios", methods=["GET", "POST"])
@login_required
@role_required("super_admin")
def users_manage():
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        role = request.form.get("role", "").strip()

        if not username or not password or role not in {"super_admin", "personal"}:
            flash("Datos de usuario invalidos.", "danger")
            return redirect(url_for("users_manage"))

        if User.query.filter_by(username=username).first():
            flash("El usuario ya existe.", "danger")
            return redirect(url_for("users_manage"))

        user = User(
            username=username,
            password_hash=generate_password_hash(password),
            role=role,
        )
        db.session.add(user)
        db.session.commit()
        flash("Usuario creado correctamente.", "success")
        return redirect(url_for("users_manage"))

    users = User.query.order_by(User.id.asc()).all()
    return render_template("users.html", users=users)


@app.route("/usuarios/<int:user_id>/generar-recuperacion", methods=["POST"])
@login_required
@role_required("super_admin")
def user_generate_recovery_code(user_id: int):
    user = db.session.get(User, user_id)
    if not user:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("users_manage"))

    code = generate_password_recovery_code()
    user.recovery_code = code
    user.recovery_expires_at = datetime.utcnow() + timedelta(minutes=30)
    db.session.commit()

    flash(
        f"Codigo de recuperacion para {user.username} (valido 30 min): {code}",
        "info",
    )
    return redirect(url_for("users_manage"))


@app.route("/usuarios/<int:user_id>/eliminar", methods=["POST"])
@login_required
@role_required("super_admin")
def user_delete(user_id: int):
    user = db.session.get(User, user_id)
    if not user:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("users_manage"))

    if user.id == session.get("user_id"):
        flash("No puedes eliminar tu propio usuario.", "danger")
        return redirect(url_for("users_manage"))

    if user.role == "super_admin":
        super_admin_count = User.query.filter_by(role="super_admin").count()
        if super_admin_count <= 1:
            flash("No puedes eliminar el ultimo super admin.", "danger")
            return redirect(url_for("users_manage"))

    if user.invoices:
        flash("No puedes eliminar un usuario con facturas registradas.", "danger")
        return redirect(url_for("users_manage"))

    db.session.delete(user)
    db.session.commit()
    flash("Usuario eliminado correctamente.", "info")
    return redirect(url_for("users_manage"))


@app.route("/recuperar-contrasena", methods=["GET", "POST"])
def password_recovery():
    code_prefill = request.args.get("codigo", "").strip()

    if request.method == "POST":
        code = request.form.get("code", "").strip()
        password = request.form.get("password", "")
        password_confirm = request.form.get("password_confirm", "")

        if not code:
            flash("Debes ingresar el codigo de recuperacion.", "danger")
            return render_template("password_recovery.html", code=code_prefill)

        if len(code) > 8 or not code.isdigit():
            flash("El codigo debe ser numerico y tener maximo 8 digitos.", "danger")
            return render_template("password_recovery.html", code=code)

        if len(password) < 8:
            flash("La contrasena debe tener al menos 8 caracteres.", "danger")
            return render_template("password_recovery.html", code=code)

        if password != password_confirm:
            flash("Las contrasenas no coinciden.", "danger")
            return render_template("password_recovery.html", code=code)

        user = validate_password_recovery_code(code)
        if not user:
            flash("Codigo invalido o vencido.", "danger")
            return render_template("password_recovery.html", code=code)

        user.password_hash = generate_password_hash(password)
        user.recovery_code = None
        user.recovery_expires_at = None
        db.session.commit()
        flash("Contrasena actualizada correctamente. Ya puedes iniciar sesion.", "success")
        return redirect(url_for("login"))

    return render_template("password_recovery.html", code=code_prefill)


with app.app_context():
    db.create_all()
    ensure_user_recovery_columns()
    seed_users()


if __name__ == "__main__":
    app.run(debug=True)
